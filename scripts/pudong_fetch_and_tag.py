#!/usr/bin/env python3
"""
浦东政策抓取 + 标签生成 + 飞书同步一体化脚本
Step 1: 增量抓取（HTTP → Excel 17列原始）
Step 2: 规则标签生成（→ Excel 19列，飞书Schema兼容）
Step 3: 飞书增量同步
Step 4: 抓取官网政策数并更新飞书统计表
PDF生成时：实时做 HTML 格式化（G字段章节结构化 + E/F/H/I字段加粗）
用法: python pudong_fetch_and_tag.py --sgin "your_sgin_here"
"""
import argparse, json, math, os, re, sys, time
from collections import Counter
from datetime import datetime
import openpyxl
import requests
from playwright.sync_api import sync_playwright

# ═══════════════════════════════════════════════════════════════
# 全局配置
# ═══════════════════════════════════════════════════════════════
DOWNLOADS   = "/Users/els/Downloads"
EXCEL_RAW   = os.path.join(DOWNLOADS, "pudong_policy_full_v2.xlsx")
EXCEL_OUT   = os.path.join(DOWNLOADS, "pudong_policy_full_v2_for_feishu.xlsx")
LOG_DIR     = "/tmp/pudong_sync_logs"
os.makedirs(LOG_DIR, exist_ok=True)

LIST_URL   = "https://pyd.pudong.gov.cn/pd-api/dataCenterXcx/special/list"
DETAIL_URL = "https://pyd.pudong.gov.cn/pd-api/dataCenterXcx/special/getInfo"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 MicroWechat/7.0.20.1781 NetType/WIFI MiniProgramEnv/Mac MacWechat/WMPF MacWechat/3.8.7(0x13080712) XWEB/16962",
    "Referer": "https://servicewechat.com/wxd5670d675e9994fd/5/page-frame.html",
    "Content-Type": "application/x-www-form-urlencoded",
    "Accept": "*/*",
    "xweb_xhr": "1"
}

# 飞书
FEISHU = {
    "app_id":     os.environ.get("FEISHU_APP_ID", "cli_a950307a10b8dcb1"),
    "app_secret": os.environ.get("FEISHU_APP_SECRET", "TFlBj160Jm4p48uZ3t4RETpL3qz1oxaj"),
    "sheet_token": "DwqqsS6TShlGhAteDf3cHRwvnHe",
    "sheet_id":   "0aad30",
}

# 飞书统计表（sheet_id: 2pLPm8）
STATS_SHEET_ID = "2pLPm8"
STATS_URL = "https://pyd.pudong.gov.cn/website/pud/policyretrieval"

# 飞书 Schema（21列）
FEISHU_SCHEMA = [
    "id", "policyName", "policyObject", "policyCondition",
    "policyContent", "paymentStandard", "contactInfo",
    "claimMethod", "amount", "amount_s",
    "applicableRegion", "leadDepartment",
    "start", "end", "zcReleaseTime",
    "行业标签", "申报主体", "政策能力", "门槛标签",
    "行业细分", "specialAbbreviat", "belongPolicy"
]

# ═══════════════════════════════════════════════════════════════
# 工具
# ═══════════════════════════════════════════════════════════════
def log(msg, br="\n"):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", end=br)
    sys.stdout.flush()

def fatal(msg):
    log(f"\n✗ 致命错误: {msg}")
    sys.exit(1)

def _parse_amount(v):
    v = str(v).strip()
    if not v or v in ("—", "待定", ""):
        return "待定"
    try:
        float(v)
        return v
    except:
        return v

# ── HTML 格式化（Excel/飞书写入前调用） ────────────────────────────────────────
def format_g(text):
    """G字段（policyContent）中文数字序号转HTML标签。"""
    if not text: return text
    def repl_ch1(m):
        return f'<br><strong>{m.group(0).rstrip()}</strong>'
    text = re.sub(
        r'([一二三四五六七八九十百]+[、．.]\s*[一-龥]{2,15}?)(?=[　\s　，,。.：:（(「"]|$)',
        repl_ch1, text
    )
    def repl_ch2(m):
        return f'<br><em>{m.group(0).rstrip()}</em>'
    text = re.sub(
        r'([（(][一二三四五六七八九十]+[）)]\s*[一-龥]{2,15}?)(?=[　\s　，,。.：:（(「"]|$)',
        repl_ch2, text
    )
    return text

def format_field(text):
    """E/F/H/I字段（政策对象/条件/兑付标准/联系方式）正则HTML格式化。"""
    if not text: return ""
    text = re.sub(
        r'([0-9一二三四五六七八九十百千万亿]+(?:\.\d+)?)([万千百亿]*)(元(?:/月|/人次|/平方米|/个)?)',
        r'<strong>\1\2</strong>\3', text
    )
    text = re.sub(r'(达标快享|免申即享|即申即享|免申直达)',
                   r'<span class="tag">\1</span>', text)
    paragraphs = re.split(r'(?<=[。！？；])', text)
    result = []
    for p in paragraphs:
        p = p.strip()
        if p: result.append(f'<p>{p}</p>')
    return "".join(result)

def col_letter(n):
    return chr(ord('A') + n)

def range_str(c0, c1, r0, r1):
    return f"{FEISHU['sheet_id']}!{col_letter(c0)}{r0}:{col_letter(c1)}{r1}"

def get_feishu_token():
    r = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU["app_id"], "app_secret": FEISHU["app_secret"]},
        timeout=10, proxies={"http": None, "https": None}
    )
    return r.json()["tenant_access_token"]

def feishu_get(url, token):
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15, proxies={"http": None, "https": None})
    # 检查响应状态码
    if r.status_code != 200:
        log(f"  ⚠️ 飞书API返回错误: HTTP {r.status_code}")
        log(f"  响应内容: {r.text[:500]}")
        return {"code": -1, "msg": f"HTTP {r.status_code}"}
    # 尝试解析 JSON
    try:
        return r.json()
    except Exception as e:
        log(f"  ⚠️ 飞书API响应解析失败: {e}")
        log(f"  响应内容: {r.text[:500]}")
        return {"code": -1, "msg": f"JSON解析失败: {e}"}

def feishu_post(url, payload, token):
    r = requests.post(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=payload, timeout=20, proxies={"http": None, "https": None}
    )
    # 检查响应状态码
    if r.status_code != 200:
        log(f"  ⚠️ 飞书API返回错误: HTTP {r.status_code}")
        log(f"  响应内容: {r.text[:500]}")
        return {"code": -1, "msg": f"HTTP {r.status_code}"}
    # 尝试解析 JSON
    try:
        return r.json()
    except Exception as e:
        log(f"  ⚠️ 飞书API响应解析失败: {e}")
        log(f"  响应内容: {r.text[:500]}")
        return {"code": -1, "msg": f"JSON解析失败: {e}"}

def strip_html(text):
    if not text:
        return ""
    return re.sub(r'<[^>]+>', '', str(text))

def strip_feishu_rich(val):
    if isinstance(val, list):
        parts = []
        for item in val:
            if isinstance(item, dict):
                if item.get("type") == "text":
                    parts.append(item.get("text", ""))
                elif item.get("type") == "url":
                    parts.append(item.get("text", ""))
            elif isinstance(item, str):
                parts.append(item)
        return "".join(parts)
    return str(val) if val else ""


# ═══════════════════════════════════════════════════════════════
# Step 1: 增量抓取
# ═══════════════════════════════════════════════════════════════
def step1_fetch(sgin):
    log("═" * 60)
    log("Step 1: 增量抓取")
    log("═" * 60)

    existing = {}
    if os.path.exists(EXCEL_RAW):
        try:
            wb = openpyxl.load_workbook(EXCEL_RAW)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    existing[str(row[0])] = row
            log(f"  已有 {len(existing)} 条记录")
        except Exception as e:
            log(f"  读已有失败: {e}，从头开始")
    else:
        log("  无已有记录，从头开始")

    log("  获取API列表...")
    api_ids = set()
    api_list_success = False
    page = 1
    while True:
        url = f"{LIST_URL}?pageNum={page}&pageSize=50&orderType=2"
        try:
            resp = requests.get(url, headers={**HEADERS, "sgin": sgin}, timeout=15).json()
        except Exception as e:
            log(f"\n  网络错误: {e}")
            break
        if resp.get("_auth_error"):
            log(f"\n  ⚠️ sgin已过期，保留已有数据")
            break
        rows = resp.get("data", {}).get("rows", [])
        if not rows:
            break
        api_list_success = True
        for row in rows:
            api_ids.add(str(row["id"]))
        log(f"  p{page}... +{len(rows)} (累计{len(api_ids)})")
        if len(rows) < 50:
            break
        page += 1
        time.sleep(0.15)

    if api_list_success:
        removed = 0
        for pid in list(existing.keys()):
            if pid not in api_ids:
                log(f"  下架: {pid}")
                del existing[pid]
                removed += 1
        if removed:
            log(f"  已删除 {removed} 条下架记录")

    to_fetch = [pid for pid in api_ids if pid not in existing]
    log(f"  待抓: {len(to_fetch)} 条")

    if not to_fetch:
        if api_list_success:
            log("  全部完成，无需抓取")
            _save_excel_17(existing, EXCEL_RAW)
        return bool(api_list_success)

    ok = fail = 0
    for i, pid in enumerate(to_fetch):
        url = f"{DETAIL_URL}?id={pid}"
        try:
            r = requests.get(url, headers={**HEADERS, "sgin": sgin}, timeout=15).json()
        except Exception as e:
            log(f"\n  网络错误: {e}")
            break
        if r.get("_auth_error"):
            log(f"\n  ⚠️ sgin过期，已抓{ok}条")
            break
        d = r.get("data", {})
        if d and d.get("specialName"):
            existing[pid] = (
                d.get("id", pid),
                d.get("specialId", ""),
                d.get("specialName", ""),
                d.get("belongPolicy", ""),
                d.get("policyObject", ""),
                d.get("policyCondition", ""),
                d.get("policyContent", ""),
                d.get("paymentStandard", ""),
                d.get("contactInfo", ""),
                d.get("claimMethod", ""),
                d.get("maxPaymentAmount", ""),
                d.get("applicableRegion", ""),
                d.get("leadDepartment", ""),
                d.get("declarStartTime", ""),
                d.get("declarEndTime", ""),
                d.get("zcReleaseTime", ""),
                d.get("specialAbbreviat", ""),
            )
            ok += 1
            log(f"  [{i+1}/{len(to_fetch)}] {pid} OK", br="")
        else:
            fail += 1
            log(f"  [{i+1}/{len(to_fetch)}] {pid} FAIL")
        if (i+1) % 5 == 0:
            _save_excel_17(existing, EXCEL_RAW)
            log(f"    已存 {len(existing)} 条", br="")
        time.sleep(0.2)

    _save_excel_17(existing, EXCEL_RAW)
    log(f"\n  完成: OK={ok} Fail={fail} Total={len(existing)}")
    return True

def _save_excel_17(data_map, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    header = [
        "id", "specialId", "policyName", "belongPolicy",
        "policyObject", "policyCondition", "policyContent",
        "paymentStandard", "contactInfo", "claimMethod",
        "maxPaymentAmount", "applicableRegion", "leadDepartment",
        "declarStartTime", "declarEndTime", "zcReleaseTime", "specialAbbreviation"
    ]
    ws.append(header)
    for row in sorted(data_map.values(), key=lambda x: str(x[0])):
        ws.append(list(row))
    wb.save(path)


# ═══════════════════════════════════════════════════════════════
# Step 2: 规则标签生成 → Excel 19列（飞书Schema）
# ═══════════════════════════════════════════════════════════════
def step2_tags():
    log("═" * 60)
    log("Step 2: 规则标签生成 → 飞书Schema")
    log("═" * 60)

    wb_in = openpyxl.load_workbook(EXCEL_RAW)
    ws_in = wb_in.active
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        fatal("Excel为空")
    data_rows = rows[1:]

    THRESHOLD_RULES = [
        ("中小微企业",    ["中小微企业","小微企业","微型企业","初创企业","小微型企业"]),
        ("高新技术企业",  ["高新技术企业","高新企业","国家高新技术企业","高新技术培育企业"]),
        ("专精特新企业", ["专精特新","专精特新中小企业","专精特新企业"]),
        ("张江区域企业", ["张江","张江区域","张江科学城","张江示范区域"]),
        ("新招引企业",   ["新招引","新引进","招商引资","招引企业"]),
        ("外资企业",     ["外资","外商","外商投资企业","港澳台"]),
        ("金融机构",     ["银行","保险","证券","基金","信托","期货","融资租赁"]),
        ("高校科研院所", ["高校","大学","科研院所","高等院校","高职院校","科研机构"]),
        ("社会组织",     ["社会组织","民非","基金会","行业协会","商会","社团","学会"]),
        ("人才",         ["高层次人才","领军人才","人才计划","人才项目","人才奖励","人才资助"]),
    ]

    def clean_text(raw):
        if not raw:
            return ""
        text = re.sub(r'<[^>]+>', ' ', str(raw))
        text = re.sub(r'[　 ]+', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    def detect_thresholds(apply_obj, suitable_obj, subject):
        obj_text = clean_text(" ".join([str(apply_obj or ""), str(suitable_obj or "")]))
        tags = []
        for tag_name, kws in THRESHOLD_RULES:
            if any(obj_text.count(kw) > 0 for kw in kws):
                tags.append(tag_name)
        if subject == "个人":
            for t in ["中小微企业","高新技术企业","专精特新企业","张江区域企业","新招引企业","外资企业"]:
                tags = [x for x in tags if x != t]
        if subject == "社会组织":
            tags = [x for x in tags if x not in {"外资企业","新招引企业"}]
        return "/".join(tags) if tags else "无限定"

    SUBJECT_RULES = [
        ("企业",     ["企业","公司","单位","集团"]),
        ("个人",     ["个人","居民","自然人","创业者"]),
        ("社会组织", ["社会组织","民非","基金会","行业协会","商会","社团","学会"]),
    ]
    def detect_subject(obj_text):
        scores = Counter()
        for tag, kws in SUBJECT_RULES:
            scores[tag] = sum(obj_text.count(kw) for kw in kws)
        if not scores:
            return "企业"
        if scores["个人"] > 0 and scores["企业"] == 0:
            return "个人"
        if scores["社会组织"] > 0 and scores["企业"] == 0:
            return "社会组织"
        return scores.most_common(1)[0][0]

    CAP_RULES = [
        ("资金补贴",  ["补贴","资助","奖励","扶持","专项资金","财政补贴","经费"]),
        ("研发支持",  ["研发","技术创新","成果转化","中试","产学研","揭榜挂帅","科技攻关"]),
        ("人才支持",  ["人才","引进","落户","住房","子女入学","医疗","人才公寓","高层次人才"]),
        ("资质认定",  ["资质","认定","评定","评审","认证","示范","试点","挂牌"]),
        ("税费减免",  ["减免","免税","退税","优惠税率","加计扣除","税收优惠"]),
        ("融资支持",  ["融资","贷款","担保","贴息","信贷","上市","股权投资","创投","基金"]),
        ("场地支持",  ["场地","办公","租金","载体","楼宇","孵化器","众创空间"]),
        ("示范推广",  ["推广","示范","展示","推介","品牌","标准制定"]),
    ]
    def detect_caps(core_text):
        scores = Counter()
        for tag, kws in CAP_RULES:
            scores[tag] = sum(core_text.count(kw) for kw in kws)
        if not scores:
            return ""
        return "/".join([t for t, _ in scores.most_common(4)])

    # ── 两级行业标签 ─────────────────────────────────────────────────────────────
    INDUSTRY_PRIMARY = [
        ("人工智能", ["人工智能","AI大模型","机器学习","深度学习","神经网络",
                       "AIGC","生成式AI","智能算法","算力"]),
        ("智能制造", ["智能制造","数字化转型","工业互联网","智能工厂","黑灯工厂",
                       "智能产线","自动化设备","数字化改造","国产化替代","高端装备",
                       "数控机床","制造智能化","人形机器人"]),
        ("生物医药", ["生物医药","创新药","生物制品","化学药","医疗器械","中药",
                       "疫苗","基因治疗","细胞治疗","CRO","CMO","生物企业","药品","生物医学"]),
        ("绿色低碳", ["绿色低碳","碳达峰","碳中和","新能源","光伏","风电","储能",
                       "氢能","碳交易","绿色制造","节能减排","环保产业","清洁能源",
                       "低碳","循环经济","可再生能源","绿色建筑","碳排放"]),
        ("消费文旅", ["文旅","旅游","文化体育","会展","商业零售","首发经济","夜间经济",
                       "文创","影视","电竞","游戏","演出","艺术品","旅游景区","酒店",
                       "餐饮","消费中心","商圈","零售业","进口商品","商贸"]),
        ("航运贸易", ["航运","港口","物流","跨境电商","国际贸易","自贸区","保税",
                       "供应链","冷链","仓储","船公司","航运服务","转口贸易","离岸贸易",
                       "数字贸易","贸易便利化","口岸","集装箱","国际海运","分拨中心"]),
        ("金融法律", ["金融","银行","保险","证券","基金","资产管理","融资担保","融资租赁",
                       "商业保理","法律服务","公证仲裁","科技金融","绿色金融","普惠金融",
                       "金融科技","金融人才","金融租赁"]),
        ("科技服务", ["研发服务","技术转移","成果转化","检验检测","知识产权","专利服务",
                       "商标服务","技术咨询","中介服务","产学研","科研项目","新型研发机构",
                       "研发机构","质量认证","标准化"]),
    ]
        # 行业细分：从正文提取具体行业/场景名词（非固定标签）
    SCENE_PATTERNS = [
        # 智慧场景
        re.compile(r'智慧\s*城市|智慧\s*政务|智慧\s*交通|智慧\s*医疗|智慧\s*教育|智慧\s*园区'),
        re.compile(r'数字\s*乡村|智慧\s*农业|智能\s*农业|数字\s*农业'),
        # AI + 垂直行业
        re.compile(r'AI\s*制药|人工智能\s*医药|AI\s*医疗|智能\s*医疗|数字\s*医疗'),
        re.compile(r'AI\s*金融|金融\s*科技|科技\s*金融|数字\s*金融|智能\s*金融'),
        re.compile(r'AI\s*文创|智能\s*文创|数字\s*文创'),
        re.compile(r'AI\s*教育|智能\s*教育|智慧\s*教育|数字\s*教育'),
        re.compile(r'AI\s*研发|AI\s*设计|智能\s*设计|智能\s*创作'),
        # 新能源/绿色
        re.compile(r'新能源\s*汽车|新能源汽车|智能\s*驾驶|自动驾驶|无人\s*驾驶'),
        re.compile(r'智能\s*网联汽车|车联网|智能\s*座舱|智能\s*车载'),
        re.compile(r'光伏\s*产业|风电\s*产业|储能\s*产业|氢能\s*产业'),
        re.compile(r'碳\s*交易|碳\s*金融|碳\s*资产|碳\s*市场|碳\s*核查'),
        re.compile(r'绿色\s*建筑|低碳\s*建筑|近零\s*碳|零碳\s*建筑'),
        re.compile(r'新能源汽车|动力电池|充电\s*桩|换电'),
        # 生物医药
        re.compile(r'创新\s*药|生物\s*药|化学\s*药|生物制品|细胞\s*治疗|基因\s*治疗'),
        re.compile(r'mRNA|ADC\s*药物|CAR-T|抗体\s*药物'),
        re.compile(r'高端\s*医疗器械|植介入|体外诊断|IVD|诊断试剂|影像设备'),
        re.compile(r'中药\s*创新|中医\s*药|中药\s*材'),
        re.compile(r'生物医药.*CDMO|CRO\s*服务|合同\s*研发'),
        # 集成电路/半导体
        re.compile(r'集成电路\s*制造|芯片\s*制造|半导体\s*制造|晶圆\s*制造'),
        re.compile(r'芯片\s*设计|集成电路\s*设计|EDA\s*软件'),
        re.compile(r'先进\s*封装|封测|SiP|CMOS'),
        re.compile(r'半导体\s*材料|光刻\s*胶|硅片|晶圆'),
        # 新兴赛道
        re.compile(r'商业\s*航天|卫星\s*互联网|运载火箭|商业\s*火箭'),
        re.compile(r'低空\s*经济|无人机|eVTOL|低空\s*物流|低空\s*出行'),
        re.compile(r'量子\s*科技|量子\s*计算|量子\s*通信|量子\s*传感'),
        re.compile(r'元宇宙|Web3|NFT|数字藏品|虚拟\s*现实|VR\s*AR'),
        re.compile(r'6G|卫星互联网|星地\s*融合|天地\s*一体化'),
        re.compile(r'合成\s*生物|生物\s*制造|生物\s*基材料'),
        re.compile(r'前沿\s*新材料|石墨烯|碳纤维|超材料|智能\s*材料'),
        # 数字经济
        re.compile(r'跨境\s*电商|数字\s*贸易|外贸\s*新业态|海外\s*仓'),
        re.compile(r'工业\s*互联网|产业\s*互联网|产业\s*数字化'),
        re.compile(r'数据\s*要素|数据\s*资产|数据\s*交易|公共\s*数据'),
        re.compile(r'云计算|边缘\s*计算|雾\s*计算|算力\s*网络'),
        re.compile(r'区块链|分布式\s*账本|可信\s*计算'),
        # 消费/文化
        re.compile(r'首店\s*经济|首发\s*经济|新品\s*发布|首店'),
        re.compile(r'夜间\s*经济|首申\s*即享|免申\s*即享'),
        re.compile(r'影视\s*产业|动漫\s*产业|电竞\s*产业|游戏\s*产业'),
        re.compile(r'文旅\s*融合|文化\s*创意|数字\s*文化|非遗'),
        # 航运/物流
        re.compile(r'国际\s*航运|航运\s*金融|船舶\s*管理|跨境\s*物流'),
        re.compile(r'冷链\s*物流|智慧\s*物流|供应链\s*金融'),
        re.compile(r'保税\s*物流|离岸\s*贸易|转口\s*贸易'),
        # 制造/装备
        re.compile(r'高端\s*装备|先进\s*制造|智能\s*装备'),
        re.compile(r'机器人|人形\s*机器人|工业\s*机器人|服务\s*机器人'),
        re.compile(r'数控\s*机床|精密\s*制造|增材\s*制造|3D\s*打印'),
        re.compile(r'航空\s*装备|航天\s*装备|海洋\s*装备'),
        # 企业服务/载体
        re.compile(r'专精特新|单项冠军|小巨人'),
        re.compile(r'孵化\s*器|众创\s*空间|加速\s*器|创业\s*载体'),
        re.compile(r'外资\s*研发|外资\s*总部|研发中心'),
        re.compile(r'高成长(?:性)?(?:青年科创)?企业|成长型科技企业'),
    ]

    def extract_industry_detail(text: str) -> str:
        """从政策正文中提取具体行业/场景词，与一级标签互补，永远有返回值。"""
        # 清理 HTML 标签
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r"\{[^}]*'type'[^}]*\}", ' ', clean)
        clean = re.sub(r"[　 \xa0]+", ' ', clean)
        clean = re.sub(r'\s+', ' ', clean).strip()

        found = []
        seen = set()
        for pattern in SCENE_PATTERNS:
            m = pattern.search(clean)
            if m:
                label = m.group(0).replace(' ', '')
                # 标准化名称（统一简称）
                std = _normalize_scene(label)
                if std and std not in seen:
                    seen.add(std)
                    found.append(std)
                    if len(found) >= 4:
                        break
        # Fallback: 精细模式没命中时，提取"XX行业/XX产业/XX领域/XX赛道"模式
        if not found:
            scene_kws = re.findall(
                r'([^　\s,，。；：:、\(\)（）]{2,6})(?:行业|产业|领域|赛道|业态|场景)',
                clean
            )
            # 去重、过滤通用词
            stopwords = {"相关","本区","该","上述","各类","其他","重点",
                         "有关","指定","特定","新型","现代","高端",
                         "大力","积极","加快","推进","战略","主导","支柱",
                         "张江科学城","特色","功能","载体","平台",
                         "重点","主体","实际","具体"}
            for kw in scene_kws:
                if kw not in stopwords and kw not in seen:
                    seen.add(kw)
                    found.append(kw)
                    if len(found) >= 3:
                        break
        return "/".join(found)

    def _normalize_scene(label: str) -> str:
        """场景名标准化映射，避免重复和歧义"""
        norm_map = {
            "智慧城市": "智慧城市", "智慧政务": "智慧政务",
            "智慧医疗": "智慧医疗", "AI制药": "AI制药",
            "AI金融": "金融科技", "金融科技": "金融科技",
            "智能驾驶": "智能驾驶", "自动驾驶": "自动驾驶",
            "新能源智能汽车": "新能源智能汽车", "新能源汽车": "新能源汽车",
            "智能网联汽车": "智能网联汽车", "车联网": "车联网",
            "碳交易": "碳交易", "碳市场": "碳市场",
            "创新药": "创新药", "生物药": "生物药",
            "高端医疗器械": "高端医疗器械",
            "集成电路制造": "集成电路制造", "芯片制造": "芯片制造",
            "芯片设计": "芯片设计",
            "商业航天": "商业航天", "低空经济": "低空经济",
            "量子科技": "量子科技", "6G": "6G",
            "跨境电商": "跨境电商", "数字贸易": "数字贸易",
            "数据要素": "数据要素",
            "人形机器人": "人形机器人", "工业机器人": "工业机器人",
            "专精特新": "专精特新",
            "孵化器": "孵化器", "众创空间": "众创空间",
            "外资研发中心": "外资研发中心",
        }
        return norm_map.get(label, label)
    IDF_WEIGHTS = {
        "企业": 0.3, "公司": 0.3, "项目": 0.5, "支持": 0.5,
        "浦东": 0.4, "新区": 0.4, "发展": 0.5, "建设": 0.5,
        "申报": 0.4, "认定": 0.5, "补贴": 0.6, "资金": 0.5,
        "AI": 1.5, "人工智能": 1.5, "芯片": 1.3, "集成电路": 1.3,
        "碳达峰": 1.5, "碳中和": 1.5, "新能源": 1.3, "智能制造": 1.3,
        "生物医药": 1.3, "医疗器械": 1.3,
    }
    def _idf_weight(kw):
        return IDF_WEIGHTS.get(kw, 1.0)

    def detect_industry(text):
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r"\{[^}]*'type'[^}]*\}", ' ', clean)
        clean = re.sub(r"[　 ]+", ' ', clean)
        clean = re.sub(r'\s+', ' ', clean).strip()

        primary_scores = Counter()
        for tag, kws in INDUSTRY_PRIMARY:
            score = 0.0
            for kw in kws:
                count = clean.count(kw)
                if count > 0:
                    score += count * _idf_weight(kw)
            primary_scores[tag] = score

        if not primary_scores:
            return ("", "")

        top_primary = primary_scores.most_common(2)
        primary_label = "/".join([t for t, _ in top_primary]) if top_primary else ""

        detail_str = extract_industry_detail(text)

        return (primary_label, detail_str)

    def process_row(apply_obj, suitable_obj, content, standard, claim):
        full = clean_text(" ".join(filter(None, [
            str(apply_obj or ""), str(suitable_obj or ""),
            str(content or ""), str(standard or "")
        ])))
        subj_text = clean_text(str(apply_obj or ""))
        subject    = detect_subject(subj_text)
        thresholds = detect_thresholds(apply_obj, suitable_obj, subject)
        primary_ind, detail_ind = detect_industry(full)
        core_cap_text = clean_text(" ".join(filter(None, [
            str(suitable_obj or ""), str(standard or ""), str(claim or "")
        ])))
        caps = detect_caps(core_cap_text)
        return subject, thresholds, primary_ind, detail_ind, caps

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet"
    ws_out.append(FEISHU_SCHEMA)

    empty_stats = {"行业标签": 0, "申报主体": 0, "政策能力": 0, "门槛标签": 0, "行业细分": 0}
    for i, row in enumerate(data_rows):
        subj, thr, primary_ind, detail_ind, cap = process_row(
            row[4], row[5], row[6], row[7], row[8]
        )
        amount_raw = row[10]
        amount_s   = _parse_amount(amount_raw)
        amount_val = amount_s if amount_s == "待定" else str(amount_raw).strip()

        ws_out.append([
            row[0], row[2],
            format_field(str(row[4] or "")),
            format_field(str(row[5] or "")),
            format_g(str(row[6] or "")),
            format_field(str(row[7] or "")),
            format_field(str(row[8] or "")),
            row[9],
            amount_val, amount_s,
            row[11], row[12], row[13], row[14], row[15],
            primary_ind, subj, cap, thr,
            detail_ind,
            row[16], row[3],
        ])

        if not primary_ind:  empty_stats["行业标签"]  += 1
        if not subj:          empty_stats["申报主体"] += 1
        if not cap:           empty_stats["政策能力"]  += 1
        if not thr:           empty_stats["门槛标签"]  += 1
        if not detail_ind:    empty_stats["行业细分"]  += 1

        if (i+1) % 100 == 0:
            log(f"  进度 {i+1}/{len(data_rows)}")

    wb_out.save(EXCEL_OUT)
    log(f"  完成: {len(data_rows)} 条 → {EXCEL_OUT}")
    log(f"  空值: 行业={empty_stats['行业标签']} 主体={empty_stats['申报主体']} "
        f"能力={empty_stats['政策能力']} 门槛={empty_stats['门槛标签']} 细分={empty_stats['行业细分']}")



def step3_sync():
    log("═" * 60)
    log("Step 3: 飞书增量同步")
    log("═" * 60)

    # 从 Excel 读取（纯文本）
    wb = openpyxl.load_workbook(EXCEL_OUT)
    ws = wb.active
    excel_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() in ("id", "状态"):
            continue
        excel_rows.append(r)
    excel_index = {str(r[0]): r for r in excel_rows}
    log(f"  Excel: {len(excel_index)} 条 ✓")

    token = get_feishu_token()

    # 表头校验
    header_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{FEISHU['sheet_token']}/values/{range_str(0, len(FEISHU_SCHEMA)-1, 1, 1)}"
    )
    hd = feishu_get(header_url, token)
    if hd.get("code") != 0:
        fatal(f"表头读取失败: {hd}")
    feishu_cols = hd["data"]["valueRange"]["values"][0]
    if feishu_cols != FEISHU_SCHEMA:
        feishu_trimmed = [c for c in feishu_cols if c is not None]
        # 情况1：飞书只有前20列（旧版），缺 zcReleaseTime（中间插入）
        if (len(feishu_trimmed) == 20
                and feishu_trimmed[14] == "行业标签"
                and FEISHU_SCHEMA[14] == "zcReleaseTime"):
            log("  ⚠️ 飞书缺 zcReleaseTime（插入位置14，迁移现有数据）...")
            insert_col = chr(ord('A') + 14)
            resp = feishu_post(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
                {"valueRanges": [{"range": f"{FEISHU['sheet_id']}!{insert_col}1:{insert_col}1", "values": [["zcReleaseTime"]]}]},
                token
            )
            if resp.get("code") != 0:
                fatal(f"插入 zcReleaseTime 失败: {resp}")
            log("  迁移 O~T → P~U...")
            read_resp = feishu_get(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values/{FEISHU['sheet_id']}!O2:T2000",
                token)
            old_rows = read_resp.get("data", {}).get("valueRange", {}).get("values") or []
            if old_rows:
                shift = feishu_post(
                    f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
                    {"valueRanges": [{"range": f"{FEISHU['sheet_id']}!P2:U{len(old_rows)+1}", "values": old_rows}]},
                    token
                )
                if shift.get("code") == 0:
                    log(f"  数据迁移成功 ({len(old_rows)} 行) ✓")
                else:
                    log(f"  ⚠️ 数据迁移失败: {shift.get('msg')}")
            feishu_cols = FEISHU_SCHEMA
            log("  表头调整完成 ✓")
        # 情况2：飞书21列，前19列对，末尾是 specialAbbreviat/belongPolicy，缺行业细分
        elif len(feishu_trimmed) == 21 and feishu_trimmed[:19] == FEISHU_SCHEMA[:19]:
            log("  ⚠️ 飞书缺 行业细分（插入位置19，末尾数据需迁移）...")
            # 1. 在T1位置插入 行业细分
            resp = feishu_post(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
                {"valueRanges": [{"range": f"{FEISHU['sheet_id']}!T1:T1", "values": [["行业细分"]]}]},
                token
            )
            if resp.get("code") != 0:
                fatal(f"插入 行业细分 失败: {resp}")
            # 2. 把末尾的 specialAbbreviat/belongPolicy 迁移到 U~V
            read_resp = feishu_get(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values/{FEISHU['sheet_id']}!T2:U2000",
                token)
            old_rows = read_resp.get("data", {}).get("valueRange", {}).get("values") or []
            if old_rows:
                shift = feishu_post(
                    f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
                    {"valueRanges": [{"range": f"{FEISHU['sheet_id']}!U2:V{len(old_rows)+1}", "values": old_rows}]},
                    token
                )
                if shift.get("code") == 0:
                    log(f"  数据迁移成功 ({len(old_rows)} 行) ✓")
                else:
                    log(f"  ⚠️ 数据迁移失败: {shift.get('msg')}")
            feishu_cols = FEISHU_SCHEMA
            log("  表头调整完成 ✓")
        # 情况3：飞书20列，前20列与 schema 一致，需要补充末尾两列
        elif feishu_trimmed == FEISHU_SCHEMA[:20]:
            missing = FEISHU_SCHEMA[20:]
            log(f"  ⚠️ 飞书缺列: {missing}，自动补上...")
            resp = feishu_post(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
                {"valueRanges": [{"range": f"{FEISHU['sheet_id']}!U1:V1", "values": [missing]}]},
                token
            )
            if resp.get("code") != 0:
                fatal(f"自动补列失败: {resp}")
            log("  补列成功 ✓")
            feishu_cols = FEISHU_SCHEMA
        else:
            log("  ⚠️ 飞书表头与预期不符")
            log(f"  飞书: {feishu_cols}")
            log(f"  预期: {FEISHU_SCHEMA}")
            fatal("表头不一致，请先在飞书中调整列顺序")
    else:
        log("  表头校验 ✓")

    # 读取飞书全量数据
    all_vals = []
    row = 2
    while True:
        end_row = min(row + 500 - 1, 2000)
        du = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{FEISHU['sheet_token']}/values/{range_str(0, len(FEISHU_SCHEMA)-1, row, end_row)}"
        )
        resp = feishu_get(du, token)
        if resp.get("code") != 0:
            break
        vals = resp["data"]["valueRange"]["values"]
        if not vals or (len(vals) == 1 and not vals[0][0]):
            break
        all_vals.extend(vals)
        if len(vals) < 500:
            break
        row = end_row + 1
        time.sleep(0.1)

    feishu_rows = []
    for v in all_vals:
        if not v or not v[0]:
            continue
        if str(v[0]) in FEISHU_SCHEMA:
            continue
        feishu_rows.append(v)
    log(f"  飞书: {len(feishu_rows)} 行 ✓")

    # 差异分析
    feishu_index = {str(row[0]): (i + 2, row) for i, row in enumerate(feishu_rows)}
    new_records, updated, deleted_ids = [], [], []

    for pid, excel_row in excel_index.items():
        if pid not in feishu_index:
            new_records.append(excel_row)
        else:
            row_num, feishu_row = feishu_index[pid]
            changed = False
            for ci in range(1, len(FEISHU_SCHEMA)):
                fv = feishu_row[ci] if ci < len(feishu_row) else ""
                ev = excel_row[ci] if (excel_row is not None and ci < len(excel_row)) else ""
                if strip_html(strip_feishu_rich(fv)) != strip_html(str(ev if ev is not None else "")):
                    changed = True
                    break
            if changed:
                updated.append((row_num, excel_row))

    excel_ids = set(excel_index.keys())
    for pid in feishu_index:
        if pid not in excel_ids:
            deleted_ids.append(pid)

    log(f"\\n  差异: 新增={len(new_records)} 更新={len(updated)} 下架={len(deleted_ids)}")

    if not new_records and not updated and not deleted_ids:
        log("  ✅ 无变动")
        return

    # 写日志
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = f"{LOG_DIR}/sync_{ts}.log"
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "新增": [str(r[0]) for r in new_records],
            "更新": [str(r[1][0]) for r in updated],
            "下架": deleted_ids,
        }, f, ensure_ascii=False, indent=2)
    log(f"  日志: {log_file}")

    # 删除下架的行（从后往前删，避免行号变化）
    if deleted_ids:
        # 按行号降序排序，从后往前删
        deleted_rows = sorted([feishu_index[did][0] for did in deleted_ids if did in feishu_index], reverse=True)
        for rn in deleted_rows:
            # 使用飞书 API 删除行
            resp = feishu_post(
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/sheets/{FEISHU['sheet_id']}/operations/batch_update",
                {
                    "requests": [
                        {
                            "deleteDimension": {
                                "majorDimension": "ROWS",
                                "startIndex": rn - 1,  # 转换为 0-based
                                "endIndex": rn  # 删除一行
                            }
                        }
                    ]
                },
                token
            )
            status = "✓" if resp.get("code") == 0 else f"✗ {resp.get('msg')}"
            log(f"  删除行{rn}: {status}")
            time.sleep(0.1)

        # 删除行后，重新读取飞书数据，更新行号映射
        log("  重新读取飞书数据...")
        all_vals = []
        row = 2
        while True:
            end_row = min(row + 500 - 1, 2000)
            du = (
                f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                f"{FEISHU['sheet_token']}/values/{range_str(0, len(FEISHU_SCHEMA)-1, row, end_row)}"
            )
            resp = feishu_get(du, token)
            if resp.get("code") != 0:
                break
            vals = resp["data"]["valueRange"]["values"]
            if not vals or (len(vals) == 1 and not vals[0][0]):
                break
            all_vals.extend(vals)
            if len(vals) < 500:
                break
            row = end_row + 1
            time.sleep(0.1)

        feishu_rows = []
        for v in all_vals:
            if not v or not v[0]:
                continue
            if str(v[0]) in FEISHU_SCHEMA:
                continue
            feishu_rows.append(v)
        feishu_index = {str(row[0]): (i + 2, row) for i, row in enumerate(feishu_rows)}
        log(f"  飞书: {len(feishu_rows)} 行 ✓")

        # 重新计算 new_records 和 updated
        new_records = []
        updated = []
        for pid, excel_row in excel_index.items():
            if pid not in feishu_index:
                new_records.append(excel_row)
            else:
                row_num, feishu_row = feishu_index[pid]
                changed = False
                for ci in range(1, len(FEISHU_SCHEMA)):
                    fv = feishu_row[ci] if ci < len(feishu_row) else ""
                    ev = excel_row[ci] if (excel_row is not None and ci < len(excel_row)) else ""
                    if strip_html(strip_feishu_rich(fv)) != strip_html(str(ev if ev is not None else "")):
                        changed = True
                        break
                if changed:
                    updated.append((row_num, excel_row))

    # 新增记录
    if new_records:
        next_row = len(feishu_rows) + 2
        values = [list(r) for r in new_records]
        end_row = next_row + len(values) - 1
        resp = feishu_post(
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
            {"valueRanges": [{"range": range_str(0, len(FEISHU_SCHEMA)-1, next_row, end_row),
                              "values": values}]}, token
        )
        if resp.get("code") == 0:
            log(f"  ✓ 新增 {len(values)} 行 (行{next_row}–{end_row})")
        else:
            log(f"  ✗ 新增失败: {resp.get('msg')}")

    # 更新记录
    for row_num, row_data in updated:
        values = [list(row_data)]
        resp = feishu_post(
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
            {"valueRanges": [{"range": range_str(0, len(FEISHU_SCHEMA)-1, row_num, row_num),
                              "values": values}]}, token
        )
        status = "✓" if resp.get("code") == 0 else f"✗ {resp.get('msg')}"
        log(f"  {status} 更新行{row_num}: {str(row_data[1] or '')[:30]}")
        time.sleep(0.1)

    log("  ✅ 同步完成")


# ═══════════════════════════════════════════════════════════════
# Step 4: 抓取官网政策数并更新飞书统计表
# ═══════════════════════════════════════════════════════════════
def step4_update_stats():
    log("═" * 60)
    log("Step 4: 抓取官网政策数并更新飞书统计表")
    log("═" * 60)

    # 用 playwright 抓取政策数
    log(f"  访问 {STATS_URL}...")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(STATS_URL, wait_until='networkidle', timeout=30000)
            time.sleep(2)

            # 查找政策数元素
            count_text = page.locator('span:has-text("查到相关内容：")').text_content()
            # 提取数字
            match = re.search(r'查到相关内容：(\d+)', count_text)
            if match:
                policy_count = int(match.group(1))
                log(f"  官网政策总数: {policy_count}")
            else:
                log(f"  ⚠️ 未找到政策数，文本: {count_text}")
                policy_count = None

            browser.close()
    except Exception as e:
        log(f"  ⚠️ 抓取失败: {e}")
        policy_count = None

    if policy_count is None:
        log("  跳过更新统计表")
        return

    # 更新飞书统计表
    token = get_feishu_token()
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    # 更新第2行（A2:D2）
    # A2: 指标 = "官网政策总数"
    # B2: 数值 = policy_count
    # C2: 更新时间 = update_time
    # D2: 来源 = STATS_URL (作为链接)
    values = [[
        "官网政策总数",
        policy_count,
        update_time,
        [{"type": "url", "text": STATS_URL, "link": STATS_URL}]
    ]]

    resp = feishu_post(
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{FEISHU['sheet_token']}/values_batch_update",
        {"valueRanges": [{"range": f"{STATS_SHEET_ID}!A2:D2", "values": values}]},
        token
    )

    if resp.get("code") == 0:
        log(f"  ✓ 统计表更新成功: {policy_count} 条")
    else:
        log(f"  ✗ 统计表更新失败: {resp.get('msg')}")


# ═══════════════════════════════════════════════════════════════
# Step 5: 生成政策 embedding 向量（本地免费，需 sentence-transformers）
# ═══════════════════════════════════════════════════════════════
def step5_generate_embeddings():
    import base64, urllib.request

    HF_TOKEN = os.environ.get("HF_TOKEN", "").strip()

    log("═" * 60)
    log("Step 5: 生成政策 embedding 向量（HuggingFace API）")
    log("═" * 60)

    if not HF_TOKEN:
        log("  ⚠️ 未设置 HF_TOKEN 环境变量，embedding 生成跳过")
        log("  请去 https://huggingface.co/settings/tokens 申请免费 Token")
        log("  export HF_TOKEN='hf_xxxxxx' 后重新运行 --embeddings")
        return

    MODEL = "BAAI/bge-base-zh-v1.5"   # 1024维，中文优化，免费
    API_URL = f"https://api-inference.huggingface.co/pipeline/feature-extraction/{MODEL}"

    wb = openpyxl.load_workbook(EXCEL_OUT)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    # 拼接政策全文文本
    policy_texts, pids = [], []
    for row in data_rows:
        text = " ".join(filter(None, [
            str(row[1] or ""),   # policyName
            str(row[4] or ""),   # policyCondition
            str(row[6] or ""),   # policyContent
        ]))
        policy_texts.append(text[:2000])
        pids.append(str(row[0]))

    total = len(policy_texts)
    embeddings: dict[str, list[float]] = {}

    # 分批调用（每批10条，避免超时）
    BATCH = 10
    for batch_start in range(0, total, BATCH):
        batch_end = min(batch_start + BATCH, total)
        batch_texts = policy_texts[batch_start:batch_end]
        batch_pids = pids[batch_start:batch_end]

        payload = {"inputs": batch_texts, "options": {"wait_for_model": True}}
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            API_URL, data=data,
            headers={
                "Authorization": f"Bearer {HF_TOKEN}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        retries, resp_text = 0, None
        while retries < 3:
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    resp_text = resp.read().decode("utf-8")
                    break
            except Exception as e:
                retries += 1
                log(f"  ⚠️ batch {batch_start}-{batch_end} 重试 ({retries}/3): {e}")
                time.sleep(3 * retries)

        if resp_text is None:
            log(f"  ⚠️ batch {batch_start}-{batch_end} 跳过（全部重试失败）")
            continue

        try:
            vectors = json.loads(resp_text)
        except Exception as e:
            log(f"  ⚠️ batch {batch_start}-{batch_end} 解析失败: {e}")
            continue

        for pid, vec in zip(batch_pids, vectors):
            embeddings[pid] = vec

        log(f"  进度 {batch_end}/{total} ({len(embeddings)} 条已生成)")
        time.sleep(1)

    if not embeddings:
        log("  ✗ 未生成任何向量，跳过保存")
        return

    dim = len(next(iter(embeddings.values())))
    out_path = os.path.join(DOWNLOADS, "policy_embeddings.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(embeddings, f, ensure_ascii=False)

    log(f"  ✓ 完成: {len(embeddings)}/{total} 条政策向量 ({dim} 维)")
    log(f"  保存至: {out_path}")
    log("  提示: 将此 JSON 上传到 Worker KV，查询时做余弦相似度匹配")
def main():
    parser = argparse.ArgumentParser(description="浦东政策抓取+标签生成+飞书同步")
    parser.add_argument("--step4-only", action="store_true", help="只跑 Step 4 更新统计表（无需 sgin）")
    parser.add_argument("--sgin", required=False, help="浦东API sgin")
    parser.add_argument("--embeddings", action="store_true",
                        help="生成政策 embedding 向量（需先运行 --sgin 完成前4步）")
    args = parser.parse_args()

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    log("=" * 60)
    log(f"浦东政策全流程 {ts}")
    log("=" * 60)

    if args.step4_only:
        step4_update_stats()
    elif args.embeddings:
        step5_generate_embeddings()
    else:
        step1_fetch(args.sgin.strip() if args.sgin else "")
        step2_tags()
        step3_sync()
        step4_update_stats()

    log("\n" + "=" * 60)
    log("🎉 全部完成！")
    log(f"Excel: {EXCEL_OUT}")
    log("=" * 60)

if __name__ == "__main__":
    main()
